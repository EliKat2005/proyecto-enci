"""
Tests para el servicio de importación de Plan de Cuentas desde Excel.
"""

import os
import tempfile
from pathlib import Path
from openpyxl import Workbook
from django.test import TestCase
from django.contrib.auth import get_user_model
from decimal import Decimal

from contabilidad.models import Empresa, EmpresaPlanCuenta, TipoCuenta, NaturalezaCuenta
from contabilidad.services_excel_import import ExcelImportService

User = get_user_model()


class ExcelImportServiceTestCase(TestCase):
    """Tests para ExcelImportService."""

    def setUp(self):
        """Configurar datos de prueba."""
        self.user = User.objects.create_user(username='testuser', password='pass')
        self.empresa = Empresa.objects.create(
            nombre='Test Company',
            owner=self.user
        )
        self.temp_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Limpiar archivos temporales."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _crear_excel(self, datos_filas: list) -> str:
        """Helper para crear un archivo Excel de prueba."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Plan de Cuentas'

        # Header
        headers = ['Código', 'Descripción', 'Tipo', 'Naturaleza', 
                   'Estado Situación', 'Es Auxiliar', 'Código Padre']
        
        # Añadir header
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Datos
        for row_num, fila in enumerate(datos_filas, 2):
            for col, valor in enumerate(fila, 1):
                ws.cell(row=row_num, column=col, value=valor)

        # Guardar
        ruta = Path(self.temp_dir) / f'test_{len(os.listdir(self.temp_dir))}.xlsx'
        wb.save(str(ruta))
        return str(ruta)

    def test_cargar_archivo_valido(self):
        """Test: Cargar archivo Excel válido."""
        ruta = self._crear_excel([
            ['1', 'Activos', 'Activo', 'Deudora', 'Si', 'No', None],
            ['1.1', 'Caja', 'Activo', 'Deudora', 'Si', 'Si', '1'],
        ])

        servicio = ExcelImportService(ruta)
        resultado = servicio.cargar_archivo()

        self.assertTrue(resultado)
        self.assertEqual(len(servicio.datos_crudos), 2)
        self.assertFalse(servicio.errores)

    def test_cargar_archivo_no_existe(self):
        """Test: Cargar archivo que no existe."""
        servicio = ExcelImportService('/ruta/no/existe/archivo.xlsx')
        resultado = servicio.cargar_archivo()

        self.assertFalse(resultado)
        self.assertTrue(any('no encontrado' in e.lower() for e in servicio.errores))

    def test_validar_codigo_duplicado(self):
        """Test: Detectar códigos duplicados."""
        ruta = self._crear_excel([
            ['1', 'Activos', 'Activo', 'Deudora', 'Si', 'No', None],
            ['1', 'Activos 2', 'Activo', 'Deudora', 'Si', 'No', None],
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, errores, _ = servicio.validar_y_corregir()

        self.assertTrue(any('duplicado' in e.lower() for e in errores))

    def test_validar_tipo_invalido(self):
        """Test: Rechazar tipo inválido."""
        ruta = self._crear_excel([
            ['1', 'Activos', 'TipoInvalido', 'Deudora', 'Si', 'No', None],
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, errores, _ = servicio.validar_y_corregir()

        self.assertTrue(any('tipo inválido' in e.lower() for e in errores))

    def test_auto_corregir_naturaleza_desde_tipo(self):
        """Test: Auto-corrección de naturaleza según tipo."""
        ruta = self._crear_excel([
            ['2', 'Pasivos', 'Pasivo', 'DeudoRA', 'Si', 'No', None],  # Deudora es incorrecta para Pasivo
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, errores, advertencias = servicio.validar_y_corregir()

        self.assertEqual(len(datos), 1)
        # Debe haber una advertencia por inconsistencia
        self.assertTrue(any('inconsistencia' in a.lower() for a in advertencias))
        # Pero la naturaleza se infiere de Pasivo (Acreedora)
        self.assertEqual(datos[0]['naturaleza'], NaturalezaCuenta.ACREEDORA)

    def test_auto_corregir_capitalizacion(self):
        """Test: Auto-corrección de capitalización."""
        ruta = self._crear_excel([
            ['1', 'activos corrientes', 'Activo', 'Deudora', 'Si', 'No', None],
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, errores, _ = servicio.validar_y_corregir()

        self.assertEqual(len(datos), 1)
        self.assertEqual(datos[0]['descripcion'], 'Activos corrientes')
        self.assertTrue(any('capitaliz' in c.lower() for c in servicio.correcciones))

    def test_validar_jerarquia_padre_no_existe(self):
        """Test: Detectar padre inexistente."""
        ruta = self._crear_excel([
            ['1.1', 'Caja', 'Activo', 'Deudora', 'Si', 'Si', '99'],  # Padre 99 no existe
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        errores_jerarquia = servicio.validar_jerarquia(datos)
        self.assertTrue(any('padre' in e.lower() and 'no existe' in e.lower() for e in errores_jerarquia))

    def test_validar_jerarquia_cuenta_auxiliar_con_hijas(self):
        """Test: Detectar cuenta auxiliar que tiene hijas."""
        ruta = self._crear_excel([
            ['1', 'Activos', 'Activo', 'Deudora', 'Si', 'Si', None],  # Auxiliar
            ['1.1', 'Caja', 'Activo', 'Deudora', 'Si', 'Si', '1'],  # Hija
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        errores_jerarquia = servicio.validar_jerarquia(datos)
        self.assertTrue(any('no puede ser auxiliar' in e.lower() for e in errores_jerarquia))

    def test_importar_cuentas_simple(self):
        """Test: Importar cuentas simples sin padre."""
        ruta = self._crear_excel([
            ['1', 'Activos', 'Activo', 'Deudora', 'Si', 'No', None],
            ['2', 'Pasivos', 'Pasivo', 'Acreedora', 'Si', 'No', None],
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        cantidad, errores = servicio.importar(self.empresa, datos)

        self.assertEqual(cantidad, 2)
        self.assertEqual(len(errores), 0)
        self.assertEqual(self.empresa.cuentas.count(), 2)

    def test_importar_cuentas_con_jerarquia(self):
        """Test: Importar cuentas con jerarquía padre-hijo."""
        ruta = self._crear_excel([
            ['1', 'Activos', 'Activo', 'Deudora', 'Si', 'No', None],
            ['1.1', 'Activo Corriente', 'Activo', 'Deudora', 'Si', 'No', '1'],
            ['1.1.01', 'Caja', 'Activo', 'Deudora', 'Si', 'Si', '1.1'],
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        cantidad, errores = servicio.importar(self.empresa, datos)

        self.assertEqual(cantidad, 3)
        self.assertEqual(len(errores), 0)

        # Verificar jerarquía
        cuenta_caja = self.empresa.cuentas.get(codigo='1.1.01')
        self.assertIsNotNone(cuenta_caja.padre)
        self.assertEqual(cuenta_caja.padre.codigo, '1.1')
        self.assertEqual(cuenta_caja.padre.padre.codigo, '1')

    def test_importar_no_afecta_cuentas_existentes(self):
        """Test: Importación no debe eliminar cuentas existentes."""
        # Crear cuenta existente
        EmpresaPlanCuenta.objects.create(
            empresa=self.empresa,
            codigo='9',
            descripcion='Cuentas Existentes',
            tipo=TipoCuenta.PATRIMONIO,
            naturaleza=NaturalezaCuenta.ACREEDORA,
            es_auxiliar=False,
            estado_situacion=True,
            activa=True
        )

        ruta = self._crear_excel([
            ['1', 'Activos', 'Activo', 'Deudora', 'Si', 'No', None],
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        cantidad, errores = servicio.importar(self.empresa, datos)

        # Debe tener 2 cuentas: la existente + la nueva
        self.assertEqual(self.empresa.cuentas.count(), 2)
        self.assertTrue(self.empresa.cuentas.filter(codigo='9').exists())

    def test_generar_reporte(self):
        """Test: Generación de reporte de validación."""
        ruta = self._crear_excel([
            ['1', 'activos', 'Activo', 'Deudora', 'Si', 'No', None],
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        reporte = servicio.generar_reporte(datos)

        self.assertIn('REPORTE', reporte)
        self.assertIn('RESUMEN', reporte)
        self.assertIn(f'Total filas', reporte)

    def test_validar_estado_situacion_inferido(self):
        """Test: Estado situación se infiere desde tipo."""
        ruta = self._crear_excel([
            ['4', 'Ingresos', 'Ingreso', 'Acreedora', None, 'No', None],  # Estado vacío
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        self.assertEqual(len(datos), 1)
        # Ingresos tienen estado_situacion = False
        self.assertFalse(datos[0]['estado_situacion'])

    def test_validar_ciclos_jerarquia(self):
        """Test: Detectar ciclos en la jerarquía."""
        # Nota: Este test requeriría referencias circulares
        # que son difíciles de crear en un archivo plano
        # pero el código debe detectarlas
        ruta = self._crear_excel([
            ['1', 'Cuenta A', 'Activo', 'Deudora', 'Si', 'No', '2'],
            ['2', 'Cuenta B', 'Activo', 'Deudora', 'Si', 'No', '1'],  # Ciclo: 1->2->1
        ])

        servicio = ExcelImportService(ruta)
        servicio.cargar_archivo()
        datos, _, _ = servicio.validar_y_corregir()

        errores_jerarquia = servicio.validar_jerarquia(datos)
        self.assertTrue(any('ciclo' in e.lower() for e in errores_jerarquia))
