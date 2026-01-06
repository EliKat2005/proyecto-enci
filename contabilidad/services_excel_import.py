"""
Servicio de importación de Plan de Cuentas desde Excel.

Incluye:
- Lectura y validación de estructura Excel
- Validaciones contables (jerarquía, naturaleza, códigos)
- Auto-corrección de errores comunes
- Importación segura a base de datos
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.core.exceptions import ValidationError
from django.db import transaction
from decimal import Decimal
from typing import Dict, List, Tuple, Optional

from .models import (
    Empresa,
    EmpresaPlanCuenta,
    TipoCuenta,
    NaturalezaCuenta
)


class ExcelImportService:
    """Servicio para importar Plan de Cuentas desde archivo Excel."""
    
    # Columnas esperadas en el Excel
    COLUMNAS_REQUERIDAS = {
        'codigo': ['Código', 'CODE', 'Cod'],
        'descripcion': ['Descripción', 'Desc', 'Description'],
        'tipo': ['Tipo', 'Type', 'TipoCuenta'],
        'naturaleza': ['Naturaleza', 'Nature', 'Deudora/Acreedora'],
        'estado_situacion': ['Estado Situación', 'Balance', 'Estado'],
        'es_auxiliar': ['Es Auxiliar', 'Auxiliar', 'Leaf'],
        'codigo_padre': ['Código Padre', 'Parent Code', 'Padre']
    }
    
    # Mapeo de naturalezas aceptadas
    NATURALEZA_MAP = {
        'deudora': NaturalezaCuenta.DEUDORA,
        'acreedora': NaturalezaCuenta.ACREEDORA,
        'D': NaturalezaCuenta.DEUDORA,
        'A': NaturalezaCuenta.ACREEDORA,
        'D.': NaturalezaCuenta.DEUDORA,
        'A.': NaturalezaCuenta.ACREEDORA,
    }
    
    # Mapeo de tipos de cuentas aceptados
    TIPO_MAP = {
        'activo': TipoCuenta.ACTIVO,
        'pasivo': TipoCuenta.PASIVO,
        'patrimonio': TipoCuenta.PATRIMONIO,
        'ingreso': TipoCuenta.INGRESO,
        'costo': TipoCuenta.COSTO,
        'gasto': TipoCuenta.GASTO,
        'asset': TipoCuenta.ACTIVO,
        'liability': TipoCuenta.PASIVO,
        'equity': TipoCuenta.PATRIMONIO,
        'revenue': TipoCuenta.INGRESO,
        'cost': TipoCuenta.COSTO,
        'expense': TipoCuenta.GASTO,
    }
    
    # Naturaleza esperada por tipo
    NATURALEZA_POR_TIPO = {
        TipoCuenta.ACTIVO: NaturalezaCuenta.DEUDORA,
        TipoCuenta.PASIVO: NaturalezaCuenta.ACREEDORA,
        TipoCuenta.PATRIMONIO: NaturalezaCuenta.ACREEDORA,
        TipoCuenta.INGRESO: NaturalezaCuenta.ACREEDORA,
        TipoCuenta.COSTO: NaturalezaCuenta.DEUDORA,
        TipoCuenta.GASTO: NaturalezaCuenta.DEUDORA,
    }
    
    # Estado situación esperado por tipo
    ESTADO_SITUACION_POR_TIPO = {
        TipoCuenta.ACTIVO: True,
        TipoCuenta.PASIVO: True,
        TipoCuenta.PATRIMONIO: True,
        TipoCuenta.INGRESO: False,
        TipoCuenta.COSTO: False,
        TipoCuenta.GASTO: False,
    }
    
    def __init__(self, ruta_archivo: str):
        """Inicializa el servicio con un archivo Excel."""
        self.ruta_archivo = ruta_archivo
        self.workbook = None
        self.worksheet = None
        self.headers = {}
        self.errores = []
        self.advertencias = []
        self.correcciones = []
        self.datos_crudos = []
    
    def cargar_archivo(self) -> bool:
        """Carga y valida el archivo Excel."""
        try:
            self.workbook = load_workbook(self.ruta_archivo)
            self.worksheet = self.workbook.active
            self._extraer_headers()
            self._extraer_datos()
            return True
        except FileNotFoundError:
            self.errores.append(f'Archivo no encontrado: {self.ruta_archivo}')
            return False
        except Exception as e:
            self.errores.append(f'Error al cargar Excel: {str(e)}')
            return False
    
    def _extraer_headers(self):
        """Extrae y mapea los headers del Excel."""
        primera_fila = list(self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        for col_num, valor in enumerate(primera_fila, 1):
            if not valor:
                continue
            
            valor_lower = str(valor).lower().strip()
            
            # Buscar coincidencia en columnas requeridas (más específicas primero)
            for campo, variaciones in self.COLUMNAS_REQUERIDAS.items():
                for variacion in variaciones:
                    # Búsqueda exacta o match exacto de palabras
                    if valor_lower == variacion.lower() or (
                        variacion.lower() in valor_lower and 
                        # Evitar falsos positivos (ej: "código padre" cuando buscamos "código")
                        not (campo == 'codigo' and 'padre' in valor_lower)
                    ):
                        self.headers[campo] = col_num
                        break
        
        # Validar que estén todas las columnas requeridas
        columnas_faltantes = set(self.COLUMNAS_REQUERIDAS.keys()) - set(self.headers.keys())
        if columnas_faltantes:
            self.errores.append(
                f'Columnas faltantes en Excel: {", ".join(columnas_faltantes)}'
            )
    
    def _extraer_datos(self):
        """Extrae los datos de cuentas del Excel."""
        # Iterar desde row 2 en adelante (row 1 es header)
        for row_num in range(2, self.worksheet.max_row + 1):
            datos = {}
            tiene_codigo = False
            
            for campo, col_num in self.headers.items():
                cell = self.worksheet.cell(row=row_num, column=col_num)
                valor = cell.value
                
                if valor is not None:
                    # Limpiar y normalizar
                    if isinstance(valor, str):
                        valor = valor.strip()
                    
                    datos[campo] = valor
                    
                    if campo == 'codigo':
                        tiene_codigo = True
            
            if tiene_codigo:  # Solo procesar filas con código
                datos['fila'] = row_num
                self.datos_crudos.append(datos)
    
    def validar_y_corregir(self) -> Tuple[List[Dict], List[str], List[str]]:
        """
        Valida y corrige los datos del Excel.
        
        Returns:
            (datos_corregidos, lista_errores, lista_advertencias)
        """
        self.errores = []
        self.advertencias = []
        self.correcciones = []
        
        datos_corregidos = []
        codigos_vistos = set()
        
        for datos in self.datos_crudos:
            fila = datos.get('fila', '?')
            errores_fila = []
            
            # 1. Validar y corregir código
            codigo = datos.get('codigo', '').strip()
            if not codigo:
                errores_fila.append(f'Fila {fila}: Código vacío')
            elif codigo in codigos_vistos:
                errores_fila.append(f'Fila {fila}: Código duplicado "{codigo}"')
            else:
                codigos_vistos.add(codigo)
            
            # 2. Validar y corregir descripción
            descripcion = datos.get('descripcion', '').strip()
            if not descripcion:
                errores_fila.append(f'Fila {fila}: Descripción vacía')
            else:
                # Auto-corrección: capitalizar primera letra
                descripcion_corr = descripcion[0].upper() + descripcion[1:] if len(descripcion) > 0 else descripcion
                if descripcion != descripcion_corr:
                    self.correcciones.append(
                        f'Fila {fila}: Descripción capitalizada: "{descripcion}" → "{descripcion_corr}"'
                    )
                    descripcion = descripcion_corr
            
            # 3. Validar y corregir tipo
            tipo_raw = str(datos.get('tipo', '')).lower().strip()
            tipo = self.TIPO_MAP.get(tipo_raw)
            if not tipo:
                errores_fila.append(f'Fila {fila}: Tipo inválido "{tipo_raw}". Válidos: {", ".join(self.TIPO_MAP.keys())}')
            
            # 4. Validar y corregir naturaleza
            naturaleza_raw = str(datos.get('naturaleza', '')).lower().strip()
            naturaleza = self.NATURALEZA_MAP.get(naturaleza_raw)
            if not naturaleza:
                # Auto-corrección: inferir desde tipo
                naturaleza_esperada = self.NATURALEZA_POR_TIPO.get(tipo)
                if naturaleza_esperada:
                    self.correcciones.append(
                        f'Fila {fila}: Naturaleza inferida desde tipo: "{naturaleza_raw}" → "{naturaleza_esperada}"'
                    )
                    naturaleza = naturaleza_esperada
                else:
                    errores_fila.append(f'Fila {fila}: Naturaleza inválida "{naturaleza_raw}". Válidas: Deudora, Acreedora')
            
            # 5. Validar consistencia tipo-naturaleza
            if tipo and naturaleza:
                naturaleza_esperada = self.NATURALEZA_POR_TIPO.get(tipo)
                if naturaleza != naturaleza_esperada:
                    self.advertencias.append(
                        f'Fila {fila}: Inconsistencia naturaleza-tipo. Tipo "{tipo}" espera "{naturaleza_esperada}", '
                        f'pero se especificó "{naturaleza}"'
                    )
                    # Auto-corrección: usar la naturaleza esperada del tipo
                    self.correcciones.append(
                        f'Fila {fila}: Naturaleza corregida desde tipo: "{naturaleza}" → "{naturaleza_esperada}"'
                    )
                    naturaleza = naturaleza_esperada
            
            # 6. Validar estado_situacion
            estado_situacion_raw = str(datos.get('estado_situacion', '')).lower().strip()
            estado_situacion = estado_situacion_raw in ['true', 's', 'si', 'yes', '1', 'verdadero']
            
            # Auto-corrección: inferir desde tipo
            if not estado_situacion_raw or estado_situacion_raw in ['', 'none']:
                estado_esperado = self.ESTADO_SITUACION_POR_TIPO.get(tipo, False)
                self.correcciones.append(
                    f'Fila {fila}: Estado Situación inferido desde tipo: {estado_esperado}'
                )
                estado_situacion = estado_esperado
            
            # 7. Validar es_auxiliar
            es_auxiliar_raw = str(datos.get('es_auxiliar', '')).lower().strip()
            es_auxiliar = es_auxiliar_raw in ['true', 's', 'si', 'yes', '1', 'verdadero']
            
            # 8. Validar código_padre (si existe)
            codigo_padre = datos.get('codigo_padre', '').strip() if datos.get('codigo_padre') else None
            
            # Agregar datos corregidos
            if not errores_fila:
                datos_corregidos.append({
                    'codigo': codigo,
                    'descripcion': descripcion,
                    'tipo': tipo,
                    'naturaleza': naturaleza,
                    'estado_situacion': estado_situacion,
                    'es_auxiliar': es_auxiliar,
                    'codigo_padre': codigo_padre,
                    'fila': fila,
                })
            else:
                for error in errores_fila:
                    self.errores.append(error)
        
        return datos_corregidos, self.errores, self.advertencias
    
    def validar_jerarquia(self, datos: List[Dict]) -> List[str]:
        """
        Valida la jerarquía de cuentas.
        
        Reglas:
        - Si tiene padre, el padre debe existir en los datos
        - No puede haber ciclos
        - Cuentas con hijas no pueden ser auxiliares
        """
        errores = []
        codigos = {d['codigo']: d for d in datos}
        visitados = set()
        
        for dato in datos:
            codigo = dato['codigo']
            codigo_padre = dato['codigo_padre']
            
            # 1. Validar que el padre existe
            if codigo_padre and codigo_padre not in codigos:
                errores.append(
                    f'Fila {dato["fila"]}: Padre "{codigo_padre}" no existe en los datos'
                )
            
            # 2. Detectar ciclos
            visitados_ciclo = set()
            actual = codigo_padre
            while actual:
                if actual in visitados_ciclo:
                    errores.append(
                        f'Fila {dato["fila"]}: Ciclo detectado en jerarquía: '
                        f'{codigo} → {actual} → ... → {actual}'
                    )
                    break
                visitados_ciclo.add(actual)
                actual = codigos.get(actual, {}).get('codigo_padre')
        
        # 3. Validar que cuentas con hijas no sean auxiliares
        hijas_por_padre = {}
        for dato in datos:
            if dato['codigo_padre']:
                if dato['codigo_padre'] not in hijas_por_padre:
                    hijas_por_padre[dato['codigo_padre']] = []
                hijas_por_padre[dato['codigo_padre']].append(dato['codigo'])
        
        for padre_codigo, hijas in hijas_por_padre.items():
            padre_data = codigos.get(padre_codigo)
            if padre_data and padre_data['es_auxiliar']:
                errores.append(
                    f'Código "{padre_codigo}": No puede ser auxiliar si tiene cuentas hijas ({", ".join(hijas)})'
                )
        
        return errores
    
    @transaction.atomic
    def importar(self, empresa: Empresa, datos: List[Dict]) -> Tuple[int, List[str]]:
        """
        Importa los datos validados a la base de datos.
        
        Returns:
            (cantidad_importada, errores_ocurridos)
        """
        errores = []
        cantidad_importada = 0
        
        # Crear mapeo de código a cuenta para referencias de padre
        cuentas_por_codigo = {}
        
        # Procesar en dos pasadas: primero sin padres, luego con padres
        datos_sin_padre = [d for d in datos if not d['codigo_padre']]
        datos_con_padre = [d for d in datos if d['codigo_padre']]
        
        # Primera pasada: crear cuentas sin padre
        for dato in datos_sin_padre:
            try:
                cuenta = EmpresaPlanCuenta.objects.create(
                    empresa=empresa,
                    codigo=dato['codigo'],
                    descripcion=dato['descripcion'],
                    tipo=dato['tipo'],
                    naturaleza=dato['naturaleza'],
                    estado_situacion=dato['estado_situacion'],
                    es_auxiliar=dato['es_auxiliar'],
                    padre=None,
                    activa=True
                )
                cuentas_por_codigo[dato['codigo']] = cuenta
                cantidad_importada += 1
            except ValidationError as e:
                errores.append(f'Error al crear cuenta {dato["codigo"]}: {str(e)}')
            except Exception as e:
                errores.append(f'Error inesperado al crear cuenta {dato["codigo"]}: {str(e)}')
        
        # Segunda pasada: crear cuentas con padre
        for dato in datos_con_padre:
            try:
                padre = cuentas_por_codigo.get(dato['codigo_padre'])
                if not padre:
                    errores.append(
                        f'Línea {dato["fila"]}: Padre "{dato["codigo_padre"]}" no existe'
                    )
                    continue
                
                cuenta = EmpresaPlanCuenta.objects.create(
                    empresa=empresa,
                    codigo=dato['codigo'],
                    descripcion=dato['descripcion'],
                    tipo=dato['tipo'],
                    naturaleza=dato['naturaleza'],
                    estado_situacion=dato['estado_situacion'],
                    es_auxiliar=dato['es_auxiliar'],
                    padre=padre,
                    activa=True
                )
                cuentas_por_codigo[dato['codigo']] = cuenta
                cantidad_importada += 1
            except ValidationError as e:
                errores.append(f'Error al crear cuenta {dato["codigo"]}: {str(e)}')
            except Exception as e:
                errores.append(f'Error inesperado al crear cuenta {dato["codigo"]}: {str(e)}')
        
        return cantidad_importada, errores
    
    def generar_reporte(self, datos_corregidos: List[Dict]) -> str:
        """Genera un reporte en formato texto de la validación."""
        reporte = []
        reporte.append('=' * 80)
        reporte.append('REPORTE DE IMPORTACIÓN - PLAN DE CUENTAS')
        reporte.append('=' * 80)
        
        if self.errores:
            reporte.append('\n❌ ERRORES (Debe corregir antes de importar):')
            for error in self.errores:
                reporte.append(f'  - {error}')
        
        if self.correcciones:
            reporte.append('\n✏️  CORRECCIONES AUTOMÁTICAS APLICADAS:')
            for corr in self.correcciones:
                reporte.append(f'  - {corr}')
        
        if self.advertencias:
            reporte.append('\n⚠️  ADVERTENCIAS (Revisar):')
            for adv in self.advertencias:
                reporte.append(f'  - {adv}')
        
        reporte.append(f'\n📊 RESUMEN:')
        reporte.append(f'  - Total filas procesadas: {len(self.datos_crudos)}')
        reporte.append(f'  - Cuentas válidas: {len(datos_corregidos)}')
        reporte.append(f'  - Errores encontrados: {len(self.errores)}')
        reporte.append(f'  - Advertencias: {len(self.advertencias)}')
        reporte.append(f'  - Correcciones aplicadas: {len(self.correcciones)}')
        
        reporte.append('\n' + '=' * 80)
        
        return '\n'.join(reporte)
