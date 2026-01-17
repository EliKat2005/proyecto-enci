"""
Servicio de exportación consolidada de Excel.
Genera un archivo Excel completo con toda la información de la empresa.
"""

import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .ml_services import MLAnalyticsService
from .models import (
    EmpresaAsiento,
    EmpresaTransaccion,
    MovimientoKardex,
    ProductoInventario,
)
from .services import EstadosFinancierosService, LibroMayorService


class ExcelExportService:
    """Servicio para generar exportaciones Excel consolidadas."""

    # Estilos predefinidos - Paleta corporativa profesional
    HEADER_FILL = PatternFill(start_color="2C5F88", end_color="2C5F88", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
    SUBTITLE_FONT = Font(size=12, bold=True, color="2C5F88")
    TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=11, color="1F4E78")

    # Colores adicionales para secciones
    ACTIVO_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    PASIVO_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    PATRIMONIO_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    INGRESO_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    EGRESO_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    BORDER_THIN = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    BORDER_MEDIUM = Border(
        left=Side(style="medium", color="5B9BD5"),
        right=Side(style="medium", color="5B9BD5"),
        top=Side(style="medium", color="5B9BD5"),
        bottom=Side(style="medium", color="5B9BD5"),
    )

    def __init__(self, empresa, fecha_inicio=None, fecha_fin=None):
        """
        Inicializa el servicio de exportación.

        Args:
            empresa: Instancia de Empresa
            fecha_inicio: Fecha de inicio para reportes (default: inicio del año)
            fecha_fin: Fecha de fin para reportes (default: hoy)
        """
        self.empresa = empresa
        hoy = date.today()
        self.fecha_inicio = fecha_inicio or date(hoy.year, 1, 1)
        self.fecha_fin = fecha_fin or hoy
        self.fecha_corte = fecha_fin or hoy

    def generar_excel_completo(self) -> bytes:
        """
        Genera un archivo Excel completo con toda la información de la empresa.

        Returns:
            bytes: Contenido del archivo Excel
        """
        wb = openpyxl.Workbook()
        # Eliminar la hoja por defecto
        wb.remove(wb.active)

        # Crear todas las hojas
        self._crear_hoja_portada(wb)
        self._crear_hoja_plan_cuentas(wb)
        self._crear_hoja_libro_diario(wb)
        self._crear_hoja_libro_mayor(wb)
        self._crear_hoja_balance_comprobacion(wb)
        self._crear_hoja_balance_general(wb)
        self._crear_hoja_estado_resultados(wb)
        self._crear_hoja_kardex(wb)
        self._crear_hoja_metricas_financieras(wb)
        self._crear_hoja_tendencias(wb)
        self._crear_hoja_top_cuentas(wb)

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _aplicar_estilo_header(self, ws, row, col_start, col_end):
        """Aplica estilo a las celdas de encabezado."""
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER_THIN

    def _aplicar_estilo_titulo(self, cell):
        """Aplica estilo de título."""
        cell.font = self.TITLE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def _aplicar_estilo_subtitulo(self, cell):
        """Aplica estilo de subtítulo."""
        cell.font = self.SUBTITLE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def _aplicar_estilo_total(self, ws, row, col_start, col_end):
        """Aplica estilo a las celdas de total."""
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.TOTAL_FILL
            cell.font = self.TOTAL_FONT
            cell.border = self.BORDER_THIN

    def _autoajustar_columnas(self, ws):
        """Ajusta el ancho de las columnas automáticamente."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _crear_hoja_portada(self, wb):
        """Crea la hoja de portada con información general."""
        ws = wb.create_sheet("Portada")

        # Título principal con fondo
        ws.merge_cells("B2:E2")
        ws["B2"] = "REPORTE CONTABLE COMPLETO"
        ws["B2"].font = Font(size=20, bold=True, color="FFFFFF")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].fill = PatternFill(start_color="2C5F88", end_color="2C5F88", fill_type="solid")
        ws.row_dimensions[2].height = 35

        # Información de la empresa con cuadro
        ws["B4"] = "INFORMACIÓN DE LA EMPRESA"
        ws["B4"].font = Font(bold=True, size=13, color="1F4E78")

        ws["B5"] = "Empresa:"
        ws["B5"].font = Font(bold=True, size=11)
        ws["C5"] = self.empresa.nombre
        ws["C5"].font = Font(size=11)

        ws["B6"] = "ID Empresa:"
        ws["B6"].font = Font(bold=True, size=10)
        ws["C6"] = f"#{self.empresa.id}"

        ws["B7"] = "Propietario:"
        ws["B7"].font = Font(bold=True, size=10)
        ws["C7"] = self.empresa.owner.get_full_name() or self.empresa.owner.username

        ws["B8"] = "Email:"
        ws["B8"].font = Font(bold=True, size=10)
        ws["C8"] = self.empresa.owner.email or "No especificado"

        # Periodo del reporte con fondo
        ws["B10"] = "PERIODO DEL REPORTE"
        ws["B10"].font = Font(bold=True, size=13, color="1F4E78")

        ws["B11"] = "Fecha Inicio:"
        ws["B11"].font = Font(bold=True)
        ws["C11"] = self.fecha_inicio.strftime("%d/%m/%Y")

        ws["B12"] = "Fecha Fin:"
        ws["B12"].font = Font(bold=True)
        ws["C12"] = self.fecha_fin.strftime("%d/%m/%Y")

        ws["B13"] = "Días del Periodo:"
        ws["B13"].font = Font(bold=True)
        dias = (self.fecha_fin - self.fecha_inicio).days + 1
        ws["C13"] = f"{dias} días"

        ws["B14"] = "Fecha de Generación:"
        ws["B14"].font = Font(bold=True)
        ws["C14"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Índice de contenidos con mejor formato
        ws["B16"] = "ÍNDICE DE CONTENIDOS"
        ws["B16"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["B16"].fill = PatternFill(start_color="2C5F88", end_color="2C5F88", fill_type="solid")
        ws["B16"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("B16:D16")
        ws.row_dimensions[16].height = 25

        contenido = [
            ("1", "Plan de Cuentas", "Estructura jerárquica completa del catálogo contable"),
            ("2", "Balance de Comprobación", "Saldos y movimientos detallados del periodo"),
            ("3", "Balance General", "Estado de situación financiera al corte"),
            ("4", "Estado de Resultados", "Ingresos, costos, gastos y utilidad del periodo"),
            ("5", "Métricas Financieras", "Ratios financieros e indicadores clave (KPIs)"),
            ("6", "Tendencias ML", "Análisis temporal con Machine Learning"),
            ("7", "Top Cuentas", "Cuentas más activas por volumen de transacciones"),
        ]

        row = 18
        for num, titulo, desc in contenido:
            ws[f"B{row}"] = num
            ws[f"B{row}"].font = Font(size=12, bold=True, color="2C5F88")
            ws[f"B{row}"].alignment = Alignment(horizontal="center")

            ws[f"C{row}"] = titulo
            ws[f"C{row}"].font = Font(size=11, bold=True)

            ws[f"D{row}"] = desc
            ws[f"D{row}"].font = Font(size=9, italic=True, color="595959")
            row += 1

        # Nota al pie
        ws[f"B{row+1}"] = "📊 Este reporte fue generado automáticamente por el Sistema ECAE"
        ws[f"B{row+1}"].font = Font(size=9, italic=True, color="7F7F7F")
        ws.merge_cells(f"B{row+1}:D{row+1}")

        # Ajustar anchos
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["E"].width = 15

    def _crear_hoja_plan_cuentas(self, wb):
        """Crea la hoja del plan de cuentas."""
        ws = wb.create_sheet("Plan de Cuentas")

        # Título
        ws["A1"] = "PLAN DE CUENTAS"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"

        # Encabezados
        headers = ["Código", "Descripción", "Tipo", "Naturaleza", "Auxiliar", "Activa"]
        ws.append([])  # Fila vacía
        row_header = ws.max_row + 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row_header, column=col, value=header)

        self._aplicar_estilo_header(ws, row_header, 1, len(headers))

        # Obtener cuentas ordenadas jerárquicamente
        cuentas = self.empresa.cuentas.all().select_related("padre").order_by("codigo")

        if not cuentas.exists():
            ws.append([])
            ws.append(["No hay cuentas registradas en el plan de cuentas."])
            ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="FF0000")
            self._autoajustar_columnas(ws)
            return

        for cuenta in cuentas:
            # Calcular nivel de indentación
            nivel = cuenta.codigo.count(".")
            descripcion = "  " * nivel + cuenta.descripcion

            ws.append(
                [
                    cuenta.codigo,
                    descripcion,
                    cuenta.get_tipo_display(),
                    cuenta.get_naturaleza_display(),
                    "Sí" if cuenta.es_auxiliar else "No",
                    "Sí" if cuenta.activa else "No",
                ]
            )

            # Aplicar color según nivel
            row = ws.max_row
            if nivel == 0:  # Elemento
                for col in range(1, 7):
                    ws.cell(row=row, column=col).font = Font(bold=True, size=11)
            elif nivel == 1:  # Grupo
                for col in range(1, 7):
                    ws.cell(row=row, column=col).font = Font(bold=True, size=10)

        self._autoajustar_columnas(ws)

    def _crear_hoja_libro_diario(self, wb):
        """Crea la hoja del Libro Diario."""
        ws = wb.create_sheet("Libro Diario")

        # Título
        ws["A1"] = "LIBRO DIARIO"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = (
            f"Periodo: {self.fecha_inicio.strftime('%d/%m/%Y')} - "
            f"{self.fecha_fin.strftime('%d/%m/%Y')}"
        )

        # Obtener asientos del periodo
        asientos = (
            EmpresaAsiento.objects.filter(
                empresa=self.empresa,
                fecha__range=[self.fecha_inicio, self.fecha_fin],
                anulado=False,
            )
            .prefetch_related("transacciones__cuenta")
            .order_by("fecha", "numero_asiento")
        )

        current_row = 5

        for asiento in asientos:
            # Encabezado del asiento
            ws.merge_cells(f"A{current_row}:G{current_row}")
            asiento_header = ws[f"A{current_row}"]
            asiento_header.value = (
                f"ASIENTO #{asiento.numero_asiento} - "
                f"{asiento.fecha.strftime('%d/%m/%Y')} - "
                f"Estado: {asiento.get_estado_display()}"
            )
            asiento_header.fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )
            asiento_header.font = Font(bold=True, size=10)
            asiento_header.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            # Concepto
            if asiento.concepto:
                ws.merge_cells(f"A{current_row}:G{current_row}")
                concepto_cell = ws[f"A{current_row}"]
                concepto_cell.value = f"Concepto: {asiento.concepto}"
                concepto_cell.font = Font(italic=True, size=9)
                current_row += 1

            # Encabezados de transacciones
            headers = ["Código", "Cuenta", "Tipo", "Tercero", "Debe", "Haber"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.BORDER_THIN
            current_row += 1

            # Transacciones del asiento
            total_debe = Decimal("0.00")
            total_haber = Decimal("0.00")

            transacciones = asiento.transacciones.all()
            for trans in transacciones:
                ws.cell(row=current_row, column=1, value=trans.cuenta.codigo)
                ws.cell(row=current_row, column=2, value=trans.cuenta.descripcion)
                ws.cell(row=current_row, column=3, value=trans.cuenta.get_tipo_display())
                ws.cell(
                    row=current_row,
                    column=4,
                    value=trans.tercero.nombre if trans.tercero else "-",
                )

                if trans.es_debito:
                    ws.cell(row=current_row, column=5, value=float(trans.monto))
                    total_debe += trans.monto
                else:
                    ws.cell(row=current_row, column=6, value=float(trans.monto))
                    total_haber += trans.monto

                # Formatear montos
                for col in [5, 6]:
                    cell = ws.cell(row=current_row, column=col)
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = self.BORDER_THIN

                # Bordes y alineación
                for col in [1, 2, 3, 4]:
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = self.BORDER_THIN

                current_row += 1

            # Totales del asiento
            ws.cell(row=current_row, column=4, value="TOTALES:")
            ws.cell(row=current_row, column=5, value=float(total_debe))
            ws.cell(row=current_row, column=6, value=float(total_haber))

            for col in [4, 5, 6]:
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True, size=10)
                cell.fill = self.TOTAL_FILL
                cell.border = self.BORDER_THIN
                if col in [5, 6]:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            current_row += 2  # Espacio entre asientos

        if not asientos.exists():
            ws["A5"] = "No hay asientos registrados en el periodo seleccionado."
            ws["A5"].font = Font(italic=True, color="FF0000")

        self._autoajustar_columnas(ws)

    def _crear_hoja_libro_mayor(self, wb):
        """Crea la hoja del Libro Mayor."""
        ws = wb.create_sheet("Libro Mayor")

        # Título
        ws["A1"] = "LIBRO MAYOR"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = (
            f"Periodo: {self.fecha_inicio.strftime('%d/%m/%Y')} - "
            f"{self.fecha_fin.strftime('%d/%m/%Y')}"
        )

        # Obtener cuentas con movimientos
        servicio_mayor = LibroMayorService(
            empresa=self.empresa, fecha_inicio=self.fecha_inicio, fecha_fin=self.fecha_fin
        )
        cuentas_mayorizadas = servicio_mayor.obtener_libro_mayor_completo()

        current_row = 5

        for cuenta_data in cuentas_mayorizadas:
            # Encabezado de la cuenta
            ws.merge_cells(f"A{current_row}:G{current_row}")
            cuenta_header = ws[f"A{current_row}"]
            cuenta_header.value = (
                f"CUENTA: {cuenta_data['cuenta'].codigo} - {cuenta_data['cuenta'].descripcion}"
            )
            cuenta_header.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cuenta_header.font = Font(color="FFFFFF", bold=True, size=11)
            cuenta_header.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            # Información de la cuenta
            ws[f"A{current_row}"] = f"Tipo: {cuenta_data['cuenta'].get_tipo_display()}"
            ws[f"C{current_row}"] = f"Naturaleza: {cuenta_data['cuenta'].get_naturaleza_display()}"
            ws[f"E{current_row}"] = f"Saldo Inicial: {cuenta_data['saldo_inicial']:,.2f}"
            current_row += 1

            # Encabezados de movimientos
            headers = ["Fecha", "Asiento", "Concepto", "Debe", "Haber", "Saldo"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.BORDER_THIN
            current_row += 1

            # Saldo inicial
            ws.cell(row=current_row, column=3, value="Saldo Inicial")
            ws.cell(row=current_row, column=6, value=float(cuenta_data["saldo_inicial"]))
            ws.cell(row=current_row, column=6).font = Font(bold=True)
            ws.cell(row=current_row, column=6).number_format = "#,##0.00"
            ws.cell(row=current_row, column=6).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            current_row += 1

            # Movimientos
            for mov in cuenta_data["movimientos"]:
                ws.cell(
                    row=current_row,
                    column=1,
                    value=mov["fecha"].strftime("%d/%m/%Y") if mov["fecha"] else "",
                )
                ws.cell(row=current_row, column=2, value=mov["numero_asiento"])
                ws.cell(row=current_row, column=3, value=mov["concepto"])
                ws.cell(row=current_row, column=4, value=float(mov["debe"]))
                ws.cell(row=current_row, column=5, value=float(mov["haber"]))
                ws.cell(row=current_row, column=6, value=float(mov["saldo"]))

                # Formatear
                for col in [4, 5, 6]:
                    cell = ws.cell(row=current_row, column=col)
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = self.BORDER_THIN

                for col in [1, 2, 3]:
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = self.BORDER_THIN

                current_row += 1

            # Totales
            ws.cell(row=current_row, column=3, value="TOTALES:")
            ws.cell(row=current_row, column=4, value=float(cuenta_data["total_debe"]))
            ws.cell(row=current_row, column=5, value=float(cuenta_data["total_haber"]))
            ws.cell(row=current_row, column=6, value=float(cuenta_data["saldo_final"]))

            for col in [3, 4, 5, 6]:
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True, size=10)
                cell.fill = self.TOTAL_FILL
                cell.border = self.BORDER_THIN
                if col in [4, 5, 6]:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            current_row += 2  # Espacio entre cuentas

        if not cuentas_mayorizadas:
            ws["A5"] = "No hay movimientos registrados en el periodo seleccionado."
            ws["A5"].font = Font(italic=True, color="FF0000")

        self._autoajustar_columnas(ws)

    def _crear_hoja_balance_comprobacion(self, wb):
        """Crea la hoja del balance de comprobación."""
        ws = wb.create_sheet("Balance de Comprobación")

        # Título
        ws["A1"] = "BALANCE DE COMPROBACIÓN"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = (
            f"Periodo: {self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')}"
        )

        # Encabezados
        headers = [
            "Código",
            "Cuenta",
            "Saldo Inicial Deudor",
            "Saldo Inicial Acreedor",
            "Debe",
            "Haber",
            "Saldo Final Deudor",
            "Saldo Final Acreedor",
        ]
        ws.append([])  # Fila vacía
        row_header = ws.max_row + 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row_header, column=col, value=header)

        self._aplicar_estilo_header(ws, row_header, 1, len(headers))

        # Datos
        cuentas = self.empresa.cuentas.filter(es_auxiliar=True).order_by("codigo")

        total_si_deudor = Decimal("0")
        total_si_acreedor = Decimal("0")
        total_debe = Decimal("0")
        total_haber = Decimal("0")
        total_sf_deudor = Decimal("0")
        total_sf_acreedor = Decimal("0")

        for cuenta in cuentas:
            saldos = LibroMayorService.calcular_saldos_cuenta(
                cuenta=cuenta, fecha_inicio=self.fecha_inicio, fecha_fin=self.fecha_fin
            )

            # Solo incluir cuentas con movimientos
            if (
                saldos["saldo_inicial"] != 0
                or saldos["debe"] != 0
                or saldos["haber"] != 0
                or saldos["saldo_final"] != 0
            ):
                si_d = saldos["saldo_inicial"] if saldos["saldo_inicial"] > 0 else Decimal("0")
                si_a = abs(saldos["saldo_inicial"]) if saldos["saldo_inicial"] < 0 else Decimal("0")
                sf_d = saldos["saldo_final"] if saldos["saldo_final"] > 0 else Decimal("0")
                sf_a = abs(saldos["saldo_final"]) if saldos["saldo_final"] < 0 else Decimal("0")

                ws.append(
                    [
                        cuenta.codigo,
                        cuenta.descripcion,
                        float(si_d),
                        float(si_a),
                        float(saldos["debe"]),
                        float(saldos["haber"]),
                        float(sf_d),
                        float(sf_a),
                    ]
                )

                # Acumular totales
                total_si_deudor += si_d
                total_si_acreedor += si_a
                total_debe += saldos["debe"]
                total_haber += saldos["haber"]
                total_sf_deudor += sf_d
                total_sf_acreedor += sf_a

        # Fila de totales
        ws.append([])
        row_total = ws.max_row + 1
        ws.append(
            [
                "",
                "TOTALES",
                float(total_si_deudor),
                float(total_si_acreedor),
                float(total_debe),
                float(total_haber),
                float(total_sf_deudor),
                float(total_sf_acreedor),
            ]
        )

        self._aplicar_estilo_total(ws, row_total, 1, len(headers))

        # Aplicar bordes a toda la tabla
        for row in ws.iter_rows(
            min_row=row_header, max_row=ws.max_row, min_col=1, max_col=len(headers)
        ):
            for cell in row:
                cell.border = self.BORDER_THIN

        # Formatear números
        for row in ws.iter_rows(min_row=row_header + 1, max_row=ws.max_row, min_col=3, max_col=8):
            for cell in row:
                if cell.value and isinstance(cell.value, int | float):
                    cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # Alinear textos
        for row in ws.iter_rows(min_row=row_header + 1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        self._autoajustar_columnas(ws)

    def _crear_hoja_balance_general(self, wb):
        """Crea la hoja del balance general."""
        ws = wb.create_sheet("Balance General")

        # Título
        ws["A1"] = "BALANCE GENERAL"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = f"Al: {self.fecha_corte.strftime('%d/%m/%Y')}"

        bg = EstadosFinancierosService.balance_general(self.empresa, self.fecha_corte)

        # Encabezados
        headers = ["Sección", "Código", "Cuenta", "Saldo"]
        ws.append([])
        row_header = ws.max_row + 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row_header, column=col, value=header)

        self._aplicar_estilo_header(ws, row_header, 1, len(headers))

        # ACTIVOS
        ws.append(["ACTIVO", "", "", ""])
        row_activo_header = ws.max_row
        ws.cell(row=row_activo_header, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_activo_header, column=1).fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        ws.merge_cells(f"A{row_activo_header}:D{row_activo_header}")
        ws.cell(row=row_activo_header, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[row_activo_header].height = 20

        for det in bg["detalle_activos"]:
            ws.append(
                ["ACTIVO", det["cuenta"].codigo, det["cuenta"].descripcion, float(det["saldo"])]
            )
            row = ws.max_row
            ws.cell(row=row, column=1).fill = self.ACTIVO_FILL

        # Total Activos
        row_total = ws.max_row + 1
        ws.append(["TOTAL ACTIVO", "", "", float(bg["activos"])])
        self._aplicar_estilo_total(ws, row_total, 1, 4)
        ws.cell(row=row_total, column=4).fill = self.SUCCESS_FILL
        ws.cell(row=row_total, column=4).font = Font(bold=True, size=11, color="375623")

        ws.append([])

        # PASIVOS
        ws.append(["PASIVO", "", "", ""])
        row_pasivo_header = ws.max_row
        ws.cell(row=row_pasivo_header, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_pasivo_header, column=1).fill = PatternFill(
            start_color="F4B084", end_color="F4B084", fill_type="solid"
        )
        ws.merge_cells(f"A{row_pasivo_header}:D{row_pasivo_header}")
        ws.cell(row=row_pasivo_header, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[row_pasivo_header].height = 20

        for det in bg["detalle_pasivos"]:
            ws.append(
                ["PASIVO", det["cuenta"].codigo, det["cuenta"].descripcion, float(det["saldo"])]
            )
            row = ws.max_row
            ws.cell(row=row, column=1).fill = self.PASIVO_FILL

        # Total Pasivos
        row_total = ws.max_row + 1
        ws.append(["TOTAL PASIVO", "", "", float(bg["pasivos"])])
        self._aplicar_estilo_total(ws, row_total, 1, 4)
        ws.cell(row=row_total, column=4).fill = PatternFill(
            start_color="FBE5D6", end_color="FBE5D6", fill_type="solid"
        )
        ws.cell(row=row_total, column=4).font = Font(bold=True, size=11, color="C65911")

        ws.append([])

        # PATRIMONIO
        ws.append(["PATRIMONIO", "", "", ""])
        row_patrim_header = ws.max_row
        ws.cell(row=row_patrim_header, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_patrim_header, column=1).fill = PatternFill(
            start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"
        )
        ws.merge_cells(f"A{row_patrim_header}:D{row_patrim_header}")
        ws.cell(row=row_patrim_header, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[row_patrim_header].height = 20

        for det in bg["detalle_patrimonio"]:
            ws.append(
                [
                    "PATRIMONIO",
                    det["cuenta"].codigo,
                    det["cuenta"].descripcion,
                    float(det["saldo"]),
                ]
            )
            row = ws.max_row
            ws.cell(row=row, column=1).fill = self.PATRIMONIO_FILL

        # Total Patrimonio
        row_total = ws.max_row + 1
        ws.append(["TOTAL PATRIMONIO", "", "", float(bg["patrimonio"])])
        self._aplicar_estilo_total(ws, row_total, 1, 4)
        ws.cell(row=row_total, column=4).fill = PatternFill(
            start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"
        )
        ws.cell(row=row_total, column=4).font = Font(bold=True, size=11, color="1F4E78")

        ws.append([])
        ws.append([])

        # Verificación del balance
        row_verif = ws.max_row + 1
        balanceado = bg["balanceado"]
        ws.append(
            [
                "VERIFICACIÓN",
                "",
                "",
                "✓ BALANCE CORRECTO" if balanceado else "✗ DESCUADRE DETECTADO",
            ]
        )
        ws.cell(row=row_verif, column=1).font = Font(bold=True, size=11, color="1F4E78")
        ws.cell(row=row_verif, column=4).font = Font(
            bold=True, size=12, color="00B050" if balanceado else "FF0000"
        )
        ws.cell(row=row_verif, column=4).fill = self.SUCCESS_FILL if balanceado else self.ALERT_FILL
        ws.merge_cells(f"A{row_verif}:C{row_verif}")
        ws.cell(row=row_verif, column=1).alignment = Alignment(horizontal="center")
        ws.row_dimensions[row_verif].height = 25

        # Ecuación contable
        ws.append([])
        ws.append(["Ecuación Contable:", "Activos = Pasivos + Patrimonio"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10, italic=True)
        ws.cell(row=ws.max_row, column=2).font = Font(size=10, italic=True)
        ws.append(
            [
                "Verificación:",
                f"${float(bg['activos']):,.2f} = ${float(bg['pasivos']):,.2f} + ${float(bg['patrimonio']):,.2f}",
            ]
        )
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
        ws.cell(row=ws.max_row, column=2).font = Font(size=10)

        # Formatear números
        for row in ws.iter_rows(min_row=row_header + 1, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                if cell.value and isinstance(cell.value, int | float):
                    cell.number_format = "#,##0.00"

        self._autoajustar_columnas(ws)

    def _crear_hoja_estado_resultados(self, wb):
        """Crea la hoja del estado de resultados."""
        ws = wb.create_sheet("Estado de Resultados")

        # Título
        ws["A1"] = "ESTADO DE RESULTADOS"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = (
            f"Periodo: {self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')}"
        )

        er = EstadosFinancierosService.estado_de_resultados(
            self.empresa, self.fecha_inicio, self.fecha_fin
        )

        # Encabezados
        headers = ["Sección", "Código", "Cuenta", "Monto"]
        ws.append([])
        row_header = ws.max_row + 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row_header, column=col, value=header)

        self._aplicar_estilo_header(ws, row_header, 1, len(headers))

        # INGRESOS
        ws.append(["INGRESOS", "", "", ""])
        row_ing_header = ws.max_row
        ws.cell(row=row_ing_header, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_ing_header, column=1).fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        ws.merge_cells(f"A{row_ing_header}:D{row_ing_header}")
        ws.cell(row=row_ing_header, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[row_ing_header].height = 20

        for det in er["detalle_ingresos"]:
            ws.append(
                ["INGRESOS", det["cuenta"].codigo, det["cuenta"].descripcion, float(det["monto"])]
            )
            ws.cell(row=ws.max_row, column=1).fill = self.INGRESO_FILL

        row_total = ws.max_row + 1
        ws.append(["TOTAL INGRESOS", "", "", float(er["ingresos"])])
        self._aplicar_estilo_total(ws, row_total, 1, 4)
        ws.cell(row=row_total, column=4).fill = self.SUCCESS_FILL
        ws.cell(row=row_total, column=4).font = Font(bold=True, size=11, color="375623")

        ws.append([])

        # COSTOS
        ws.append(["COSTOS", "", "", ""])
        row_costo_header = ws.max_row
        ws.cell(row=row_costo_header, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_costo_header, column=1).fill = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )
        ws.merge_cells(f"A{row_costo_header}:D{row_costo_header}")
        ws.cell(row=row_costo_header, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[row_costo_header].height = 20

        for det in er["detalle_costos"]:
            ws.append(
                ["COSTOS", det["cuenta"].codigo, det["cuenta"].descripcion, float(det["monto"])]
            )
            ws.cell(row=ws.max_row, column=1).fill = self.EGRESO_FILL

        row_total = ws.max_row + 1
        ws.append(["TOTAL COSTOS", "", "", float(er["costos"])])
        self._aplicar_estilo_total(ws, row_total, 1, 4)
        ws.cell(row=row_total, column=4).fill = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )
        ws.cell(row=row_total, column=4).font = Font(bold=True, size=11, color="C65911")

        ws.append([])

        # UTILIDAD BRUTA
        row_util_bruta = ws.max_row + 1
        ws.append(["UTILIDAD BRUTA", "", "", float(er["utilidad_bruta"])])
        ws.cell(row=row_util_bruta, column=1).font = Font(bold=True, size=11, color="1F4E78")
        ws.cell(row=row_util_bruta, column=4).font = Font(bold=True, size=11)

        ws.append([])

        # GASTOS
        ws.append(["GASTOS", "", "", ""])
        row_gasto_header = ws.max_row
        ws.cell(row=row_gasto_header, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_gasto_header, column=1).fill = PatternFill(
            start_color="E74C3C", end_color="E74C3C", fill_type="solid"
        )
        ws.merge_cells(f"A{row_gasto_header}:D{row_gasto_header}")
        ws.cell(row=row_gasto_header, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[row_gasto_header].height = 20

        for det in er["detalle_gastos"]:
            ws.append(
                ["GASTOS", det["cuenta"].codigo, det["cuenta"].descripcion, float(det["monto"])]
            )
            ws.cell(row=ws.max_row, column=1).fill = self.EGRESO_FILL

        row_total = ws.max_row + 1
        ws.append(["TOTAL GASTOS", "", "", float(er["gastos"])])
        self._aplicar_estilo_total(ws, row_total, 1, 4)
        ws.cell(row=row_total, column=4).fill = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )
        ws.cell(row=row_total, column=4).font = Font(bold=True, size=11, color="C65911")

        ws.append([])
        ws.append([])

        # UTILIDAD NETA
        row_util_neta = ws.max_row + 1
        utilidad_neta_valor = float(er["utilidad_neta"])
        ws.append(["UTILIDAD NETA", "", "", utilidad_neta_valor])
        ws.cell(row=row_util_neta, column=1).font = Font(bold=True, size=12, color="1F4E78")

        # Color según si es ganancia o pérdida
        color_utilidad = "00B050" if utilidad_neta_valor >= 0 else "FF0000"
        fill_utilidad = self.SUCCESS_FILL if utilidad_neta_valor >= 0 else self.ALERT_FILL

        ws.cell(row=row_util_neta, column=4).font = Font(bold=True, size=12, color=color_utilidad)
        ws.cell(row=row_util_neta, column=4).fill = fill_utilidad
        ws.merge_cells(f"A{row_util_neta}:C{row_util_neta}")
        ws.cell(row=row_util_neta, column=1).alignment = Alignment(horizontal="center")
        ws.row_dimensions[row_util_neta].height = 25

        # Formatear números
        for row in ws.iter_rows(min_row=row_header + 1, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                if cell.value and isinstance(cell.value, int | float):
                    cell.number_format = "#,##0.00"

        self._autoajustar_columnas(ws)

    def _crear_hoja_kardex(self, wb):
        """Crea la hoja del Kardex de Inventario."""
        ws = wb.create_sheet("Kardex")

        # Título
        ws["A1"] = "KARDEX DE INVENTARIO"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = (
            f"Periodo: {self.fecha_inicio.strftime('%d/%m/%Y')} - "
            f"{self.fecha_fin.strftime('%d/%m/%Y')}"
        )

        # Obtener productos con movimientos
        productos = ProductoInventario.objects.filter(empresa=self.empresa, activo=True).order_by(
            "codigo"
        )

        current_row = 5

        for producto in productos:
            # Encabezado del producto
            ws.merge_cells(f"A{current_row}:J{current_row}")
            producto_header = ws[f"A{current_row}"]
            producto_header.value = (
                f"PRODUCTO: {producto.codigo} - {producto.nombre} | "
                f"Unidad: {producto.unidad_medida} | "
                f"Método: {producto.get_metodo_valuacion_display()}"
            )
            producto_header.fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )
            producto_header.font = Font(color="FFFFFF", bold=True, size=11)
            producto_header.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            # Información adicional del producto
            ws[f"A{current_row}"] = f"Stock Actual: {producto.cantidad_actual}"
            ws[f"C{current_row}"] = f"Costo Promedio: ${producto.costo_promedio:,.2f}"
            ws[f"F{current_row}"] = f"Valor Total: ${producto.valor_total:,.2f}"
            current_row += 1

            # Encabezados de movimientos
            headers = [
                "Fecha",
                "Tipo",
                "Asiento",
                "Concepto",
                "Cant. Entrada",
                "$ Entrada",
                "Cant. Salida",
                "$ Salida",
                "Saldo Cant.",
                "Saldo $",
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.BORDER_THIN
            current_row += 1

            # Obtener movimientos del periodo
            movimientos = (
                MovimientoKardex.objects.filter(
                    producto=producto,
                    fecha__range=[self.fecha_inicio, self.fecha_fin],
                    anulado=False,
                )
                .select_related("asiento")
                .order_by("fecha", "created_at")
            )

            # Calcular saldo inicial (movimientos anteriores al periodo)
            movimientos_anteriores = MovimientoKardex.objects.filter(
                producto=producto, fecha__lt=self.fecha_inicio, anulado=False
            )

            cantidad_inicial = 0
            valor_inicial = Decimal("0.00")
            for mov_ant in movimientos_anteriores:
                if mov_ant.tipo_movimiento in ["CP", "AI"]:  # Entrada
                    cantidad_inicial += mov_ant.cantidad
                    valor_inicial += mov_ant.valor_total
                else:  # Salida
                    cantidad_inicial -= mov_ant.cantidad
                    valor_inicial -= mov_ant.valor_total

            # Fila de saldo inicial
            ws.cell(row=current_row, column=4, value="Saldo Inicial")
            ws.cell(row=current_row, column=9, value=cantidad_inicial)
            ws.cell(row=current_row, column=10, value=float(valor_inicial))
            ws.cell(row=current_row, column=4).font = Font(bold=True, italic=True)
            ws.cell(row=current_row, column=9).font = Font(bold=True)
            ws.cell(row=current_row, column=10).font = Font(bold=True)
            ws.cell(row=current_row, column=9).number_format = "#,##0"
            ws.cell(row=current_row, column=10).number_format = "#,##0.00"
            current_row += 1

            # Movimientos del periodo
            saldo_cantidad = cantidad_inicial
            saldo_valor = valor_inicial

            for mov in movimientos:
                # Calcular entrada/salida
                if mov.tipo_movimiento in ["CP", "AI"]:  # Entrada
                    cant_entrada = mov.cantidad
                    val_entrada = float(mov.valor_total)
                    cant_salida = 0
                    val_salida = 0
                    saldo_cantidad += mov.cantidad
                    saldo_valor += mov.valor_total
                else:  # Salida
                    cant_entrada = 0
                    val_entrada = 0
                    cant_salida = mov.cantidad
                    val_salida = float(mov.valor_total)
                    saldo_cantidad -= mov.cantidad
                    saldo_valor -= mov.valor_total

                ws.cell(row=current_row, column=1, value=mov.fecha.strftime("%d/%m/%Y"))
                ws.cell(row=current_row, column=2, value=mov.get_tipo_movimiento_display())
                ws.cell(
                    row=current_row,
                    column=3,
                    value=mov.asiento.numero_asiento if mov.asiento else "-",
                )
                ws.cell(row=current_row, column=4, value=mov.concepto or "-")
                ws.cell(row=current_row, column=5, value=cant_entrada)
                ws.cell(row=current_row, column=6, value=val_entrada)
                ws.cell(row=current_row, column=7, value=cant_salida)
                ws.cell(row=current_row, column=8, value=val_salida)
                ws.cell(row=current_row, column=9, value=saldo_cantidad)
                ws.cell(row=current_row, column=10, value=float(saldo_valor))

                # Formatear cantidades
                for col in [5, 7, 9]:
                    cell = ws.cell(row=current_row, column=col)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = self.BORDER_THIN

                # Formatear valores monetarios
                for col in [6, 8, 10]:
                    cell = ws.cell(row=current_row, column=col)
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = self.BORDER_THIN

                # Formatear textos
                for col in [1, 2, 3, 4]:
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = self.BORDER_THIN

                # Destacar tipo de movimiento con color
                tipo_cell = ws.cell(row=current_row, column=2)
                if mov.tipo_movimiento in ["CP", "AI"]:
                    tipo_cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                else:
                    tipo_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

                current_row += 1

            # Fila de saldo final
            ws.cell(row=current_row, column=4, value="Saldo Final")
            ws.cell(row=current_row, column=9, value=saldo_cantidad)
            ws.cell(row=current_row, column=10, value=float(saldo_valor))

            for col in [4, 9, 10]:
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True, size=10)
                cell.fill = self.TOTAL_FILL
                cell.border = self.BORDER_THIN
                if col in [9, 10]:
                    cell.number_format = "#,##0.00" if col == 10 else "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            current_row += 2  # Espacio entre productos

            # Mensaje si no hay movimientos
            if not movimientos.exists() and cantidad_inicial == 0:
                ws[f"A{current_row}"] = "Sin movimientos en el periodo"
                ws[f"A{current_row}"].font = Font(italic=True, color="808080")
                current_row += 1

        if not productos.exists():
            ws["A5"] = "No hay productos de inventario registrados."
            ws["A5"].font = Font(italic=True, color="FF0000")

        self._autoajustar_columnas(ws)

    def _crear_hoja_metricas_financieras(self, wb):
        """Crea la hoja de métricas financieras (ML)."""
        ws = wb.create_sheet("Métricas Financieras")

        # Título
        ws["A1"] = "MÉTRICAS FINANCIERAS Y RATIOS"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = "Generado con Machine Learning / Analytics"

        try:
            ml_service = MLAnalyticsService(self.empresa)
            metrics = ml_service.get_dashboard_metrics()

            # Métricas principales
            ws.append([])
            ws["A5"] = "INDICADORES PRINCIPALES"
            ws["A5"].font = self.SUBTITLE_FONT

            ws.append([])
            ws.append(["Indicador", "Valor", "Interpretación"])
            self._aplicar_estilo_header(ws, 7, 1, 3)

            metricas_data = [
                ("Liquidez Corriente", f"{metrics['liquidez']:.2f}", "Capacidad de pago"),
                ("ROA (Rentabilidad Activos)", f"{metrics['roa']:.2f}%", "Eficiencia activos"),
                (
                    "Endeudamiento",
                    f"{metrics['endeudamiento']:.2f}%",
                    "Nivel de endeudamiento",
                ),
                ("Margen Neto", f"{metrics['margen_neto']:.2f}%", "Rentabilidad neta"),
            ]

            for metrica in metricas_data:
                ws.append(metrica)

            ws.append([])
            ws["A13"] = "COMPOSICIÓN PATRIMONIAL"
            ws["A13"].font = self.SUBTITLE_FONT

            ws.append([])
            ws.append(["Concepto", "Monto", "Porcentaje"])
            self._aplicar_estilo_header(ws, 15, 1, 3)

            total_activos_abs = abs(metrics["activos"])
            total_base = total_activos_abs if total_activos_abs > 0 else 1

            composicion = [
                (
                    "Activos",
                    f"${metrics['activos']:,.2f}",
                    f"{(abs(metrics['activos']) / total_base * 100):.2f}%",
                ),
                (
                    "Pasivos",
                    f"${metrics['pasivos']:,.2f}",
                    f"{(abs(metrics['pasivos']) / total_base * 100):.2f}%",
                ),
                (
                    "Patrimonio",
                    f"${metrics['patrimonio']:,.2f}",
                    f"{(abs(metrics['patrimonio']) / total_base * 100):.2f}%",
                ),
            ]

            for comp in composicion:
                ws.append(comp)

            ws.append([])
            ws["A20"] = "RESULTADOS DEL PERIODO"
            ws["A20"].font = self.SUBTITLE_FONT

            ws.append([])
            ws.append(["Concepto", "Monto"])
            self._aplicar_estilo_header(ws, 22, 1, 2)

            resultados = [
                ("Ingresos Totales", f"${metrics['ingresos']:,.2f}"),
                ("Costos Totales", f"${metrics['costos']:,.2f}"),
                ("Gastos Totales", f"${metrics['gastos']:,.2f}"),
                ("Utilidad Neta", f"${metrics['utilidad_neta']:,.2f}"),
            ]

            for res in resultados:
                ws.append(res)

        except Exception as e:
            ws.append([])
            ws.append([f"Error al calcular métricas: {str(e)}"])

        self._autoajustar_columnas(ws)

    def _crear_hoja_tendencias(self, wb):
        """Crea la hoja de análisis de tendencias (ML)."""
        ws = wb.create_sheet("Tendencias ML")

        # Título
        ws["A1"] = "ANÁLISIS DE TENDENCIAS (Machine Learning)"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = "Análisis de series temporales - Últimos 12 meses"

        try:
            ml_service = MLAnalyticsService(self.empresa)
            data = ml_service.get_analytics_time_series(meses=12)

            # Encabezados
            ws.append([])
            headers = [
                "Periodo",
                "Ingresos",
                "Gastos",
                "Costos",
                "Utilidad",
                "Activos",
                "Pasivos",
                "Flujo Neto",
            ]
            ws.append(headers)
            row_header = ws.max_row
            self._aplicar_estilo_header(ws, row_header, 1, len(headers))

            if "series" in data and len(data["series"]) > 0:
                total_ingresos = 0
                total_gastos = 0
                total_costos = 0
                total_utilidad = 0

                for periodo_data in data["series"]:
                    ingresos = float(periodo_data.get("ingresos", 0))
                    gastos = float(periodo_data.get("gastos", 0))
                    costos = float(periodo_data.get("costos", 0))
                    utilidad = float(periodo_data.get("utilidad", 0))
                    activos = float(periodo_data.get("activos", 0))
                    pasivos = float(periodo_data.get("pasivos", 0))
                    flujo_neto = ingresos - gastos - costos

                    # Convertir periodo YYYY-MM a formato legible
                    periodo = periodo_data.get("periodo", "")
                    if periodo:
                        try:
                            fecha = datetime.strptime(periodo, "%Y-%m")
                            periodo_formatted = fecha.strftime("%b %Y")
                        except:
                            periodo_formatted = periodo
                    else:
                        periodo_formatted = "N/A"

                    ws.append(
                        [
                            periodo_formatted,
                            ingresos,
                            gastos,
                            costos,
                            utilidad,
                            activos,
                            pasivos,
                            flujo_neto,
                        ]
                    )

                    # Color según flujo neto
                    row = ws.max_row
                    if flujo_neto > 0:
                        ws.cell(row=row, column=8).fill = self.SUCCESS_FILL
                    elif flujo_neto < 0:
                        ws.cell(row=row, column=8).fill = self.ALERT_FILL

                    total_ingresos += ingresos
                    total_gastos += gastos
                    total_costos += costos
                    total_utilidad += utilidad

                # Fila de totales
                ws.append([])
                row_total = ws.max_row + 1
                ws.append(
                    [
                        "TOTALES",
                        total_ingresos,
                        total_gastos,
                        total_costos,
                        total_utilidad,
                        "",
                        "",
                        total_ingresos - total_gastos - total_costos,
                    ]
                )
                self._aplicar_estilo_total(ws, row_total, 1, len(headers))

                # Estadísticas adicionales
                ws.append([])
                ws.append([])
                ws.append(["ESTADÍSTICAS DEL PERIODO"])
                ws.cell(row=ws.max_row, column=1).font = self.SUBTITLE_FONT

                num_periodos = len(data["series"])
                promedio_ingresos = total_ingresos / num_periodos if num_periodos > 0 else 0
                promedio_gastos = total_gastos / num_periodos if num_periodos > 0 else 0
                promedio_utilidad = total_utilidad / num_periodos if num_periodos > 0 else 0

                ws.append(["Total de periodos analizados:", num_periodos])
                ws.append(["Promedio mensual de ingresos:", promedio_ingresos])
                ws.append(["Promedio mensual de gastos:", promedio_gastos])
                ws.append(["Promedio mensual de utilidad:", promedio_utilidad])

                # Formatear estadísticas
                for row in range(ws.max_row - 3, ws.max_row + 1):
                    ws.cell(row=row, column=1).font = Font(bold=True, size=10)
                    if ws.cell(row=row, column=2).value and isinstance(
                        ws.cell(row=row, column=2).value, int | float
                    ):
                        ws.cell(row=row, column=2).number_format = "#,##0.00"

                # Formatear números en la tabla
                for row in ws.iter_rows(
                    min_row=row_header + 1, max_row=row_total, min_col=2, max_col=8
                ):
                    for cell in row:
                        if cell.value and isinstance(cell.value, int | float):
                            cell.number_format = "#,##0.00"
                            cell.alignment = Alignment(horizontal="right", vertical="center")

                # Aplicar bordes
                for row in ws.iter_rows(
                    min_row=row_header, max_row=row_total, min_col=1, max_col=len(headers)
                ):
                    for cell in row:
                        cell.border = self.BORDER_THIN
            else:
                ws.append([])
                ws.append(["No hay datos suficientes para generar el análisis de tendencias."])
                ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="FF0000")
                ws.append(["Se requieren transacciones contables en los últimos 12 meses."])
                ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=10)

        except Exception as e:
            ws.append([])
            ws.append([f"Error al generar tendencias: {str(e)}"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF0000")
            ws.append(["Por favor, verifique que existan transacciones registradas."])
            ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=10)

        # Ajustar anchos de columnas manualmente para esta hoja (números grandes)
        ws.column_dimensions["A"].width = 12  # Periodo
        ws.column_dimensions["B"].width = 16  # Ingresos
        ws.column_dimensions["C"].width = 16  # Gastos
        ws.column_dimensions["D"].width = 16  # Costos
        ws.column_dimensions["E"].width = 16  # Utilidad
        ws.column_dimensions["F"].width = 16  # Activos
        ws.column_dimensions["G"].width = 16  # Pasivos
        ws.column_dimensions["H"].width = 16  # Flujo Neto

    def _crear_hoja_top_cuentas(self, wb):
        """Crea la hoja de top cuentas más activas."""
        ws = wb.create_sheet("Top Cuentas")

        # Título
        ws["A1"] = "TOP CUENTAS MÁS ACTIVAS"
        self._aplicar_estilo_titulo(ws["A1"])
        ws["A2"] = f"Empresa: {self.empresa.nombre}"
        ws["A3"] = (
            f"Periodo: {self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')}"
        )

        # Obtener transacciones y agrupar por cuenta
        from django.db.models import Count, Sum

        from .models import EstadoAsiento

        cuentas_activas = (
            EmpresaTransaccion.objects.filter(
                asiento__empresa=self.empresa,
                asiento__fecha__gte=self.fecha_inicio,
                asiento__fecha__lte=self.fecha_fin,
                asiento__estado=EstadoAsiento.CONFIRMADO,
            )
            .values("cuenta__codigo", "cuenta__descripcion", "cuenta__tipo")
            .annotate(
                num_transacciones=Count("id"),
                total_debe=Sum("debe"),
                total_haber=Sum("haber"),
            )
            .order_by("-num_transacciones")[:20]
        )

        ws.append([])
        ws.append(
            ["Ranking", "Código", "Cuenta", "Tipo", "# Transacciones", "Total Debe", "Total Haber"]
        )
        row_header = 5
        self._aplicar_estilo_header(ws, row_header, 1, 7)

        if cuentas_activas:
            for idx, cuenta in enumerate(cuentas_activas, start=1):
                ws.append(
                    [
                        idx,
                        cuenta["cuenta__codigo"],
                        cuenta["cuenta__descripcion"],
                        cuenta["cuenta__tipo"],
                        cuenta["num_transacciones"],
                        float(cuenta["total_debe"] or 0),
                        float(cuenta["total_haber"] or 0),
                    ]
                )

                # Destacar el top 3
                row = ws.max_row
                if idx <= 3:
                    ws.cell(row=row, column=1).fill = PatternFill(
                        start_color="FFD966", end_color="FFD966", fill_type="solid"
                    )
                    ws.cell(row=row, column=1).font = Font(bold=True, size=11)

            # Aplicar bordes
            for row in ws.iter_rows(min_row=row_header, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = self.BORDER_THIN

            # Formatear números
            for row in ws.iter_rows(
                min_row=row_header + 1, max_row=ws.max_row, min_col=5, max_col=7
            ):
                for cell in row:
                    if cell.value and isinstance(cell.value, int | float):
                        if cell.column <= 5:
                            cell.number_format = "#,##0"
                        else:
                            cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            # Alinear textos
            for row in ws.iter_rows(
                min_row=row_header + 1, max_row=ws.max_row, min_col=1, max_col=4
            ):
                for cell in row:
                    if cell.column == 1:  # Ranking centrado
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            ws.append([])
            ws.append(["No hay transacciones registradas en el periodo seleccionado."])
            ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="FF0000")

        self._autoajustar_columnas(ws)
